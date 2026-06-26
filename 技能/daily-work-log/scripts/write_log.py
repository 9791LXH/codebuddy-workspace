#!/usr/bin/env python3
"""
每日工作重点录入脚本
用法: python write_log.py <日期> <维度> <具体内容> <处理过程> [结果成效] [个人总结]
日期格式: YYYY-MM-DD 或 直接传入 Excel 序列号
"""

import sys
import openpyxl
from datetime import date, datetime

EXCEL_FILE = "E:/会议纪要/每日重点回顾--陆新海.xlsx"

# Excel 日期序列号转换（以1900-01-01为基准，但Excel有个1900年闰年bug，所以用1899-12-30）
def date_to_excel_serial(d: date) -> int:
    epoch = date(1899, 12, 30)
    return (d - epoch).days

def excel_serial_to_date(serial: int) -> date:
    epoch = date(1899, 12, 30)
    from datetime import timedelta
    return epoch + timedelta(days=serial)

def get_sheet_for_date(wb, target_date: date):
    """根据日期找到对应的工作表（按月份匹配）"""
    month = target_date.month
    month_names = {
        1: ['1月份', '第1月', 'Jan'],
        2: ['2月份', '第2月', 'Feb'],
        3: ['3月份', '第3月', 'Mar'],
        4: ['4月份', '第4月', 'Apr'],
        5: ['5月份', '第5月', 'May'],
        6: ['6月份', '6月份 ', 'Jun'],
        7: ['7月份', '第7月', 'Jul'],
        8: ['8月份', '第8月', 'Aug'],
        9: ['9月份', '第9月', 'Sep'],
        10: ['10月份', '第10月', 'Oct'],
        11: ['11月份', '第11月', 'Nov'],
        12: ['12月份', '第12月', 'Dec'],
    }
    candidates = month_names.get(month, [])
    for name in candidates:
        if name.strip() in [s.strip() for s in wb.sheetnames]:
            for s in wb.sheetnames:
                if s.strip() == name.strip():
                    return wb[s]
    # fallback: 返回最后一个工作表
    return wb.worksheets[-1]

def find_insert_position(ws, serial: int) -> int:
    """
    找到插入行位置：
    - 新记录紧跟同日期最后一行的下方
    - 如果该日期不存在，则找到前一个日期最后一行的下方
    - 如果都不存在，追加到末尾
    """
    last_row_same_date = 0
    last_row_before_date = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        val = row[0].value
        if val is None:
            continue
        if isinstance(val, datetime):
            val = date_to_excel_serial(val.date())
        if isinstance(val, date):
            val = date_to_excel_serial(val)
        try:
            val = int(val)
        except (TypeError, ValueError):
            continue
        if val == serial:
            last_row_same_date = max(last_row_same_date, row[0].row)
        elif val < serial:
            last_row_before_date = max(last_row_before_date, row[0].row)

    if last_row_same_date > 0:
        return last_row_same_date + 1
    elif last_row_before_date > 0:
        return last_row_before_date + 1
    else:
        return ws.max_row + 1

def write_log(date_str, dimension, content, process, result="", summary=""):
    # 解析日期
    try:
        if "-" in str(date_str):
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        else:
            target_date = excel_serial_to_date(int(date_str))
    except Exception as e:
        print(f"日期解析失败: {e}")
        sys.exit(1)

    serial = date_to_excel_serial(target_date)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = get_sheet_for_date(wb, target_date)

    insert_row = find_insert_position(ws, serial)
    ws.insert_rows(insert_row)

    date_cell = ws.cell(row=insert_row, column=1, value=serial)
    date_cell.number_format = 'yyyy-mm-dd;@'
    ws.cell(row=insert_row, column=2, value=dimension)
    ws.cell(row=insert_row, column=3, value=content)
    ws.cell(row=insert_row, column=4, value=process)
    ws.cell(row=insert_row, column=5, value=result)
    ws.cell(row=insert_row, column=6, value=summary)

    wb.save(EXCEL_FILE)
    print(f"写入成功: {target_date} | {dimension} | {content} -> 第{insert_row}行 (工作表: {ws.title})")

if __name__ == "__main__":
    if len(sys.argv) < 5:
        print("用法: python write_log.py <日期YYYY-MM-DD> <维度> <具体内容> <处理过程> [结果成效] [个人总结]")
        sys.exit(1)
    write_log(
        sys.argv[1],
        sys.argv[2],
        sys.argv[3],
        sys.argv[4],
        sys.argv[5] if len(sys.argv) > 5 else "",
        sys.argv[6] if len(sys.argv) > 6 else "",
    )
